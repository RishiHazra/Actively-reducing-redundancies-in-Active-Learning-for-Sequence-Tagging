import os
import tensorflow as tf
import xlsxwriter
from openpyxl import load_workbook

class BaseModel(object):
    """Generic class for general methods that are not specific to NER"""

    def __init__(self, config):
        """Defines self.config and self.logger
        Args:
            config: (Config instance) class with hyper parameters,
                vocab and embeddings
        """
        self.config = config
        self.logger = config.logger
        self.sess   = None
        self.saver  = None


    def reinitialize_weights(self, scope_name):
        """Reinitializes the weights of a given layer"""
        variables = tf.contrib.framework.get_variables(scope_name)
        init = tf.variables_initializer(variables)
        self.sess.run(init)


    def add_train_op(self, lr_method, lr, loss, clip=-1):
        """Defines self.train_op that performs an update on a batch
        Args:
            lr_method: (string) sgd method, for example "adam"
            lr: (tf.placeholder) tf.float32, learning rate
            loss: (tensor) tf.float32 loss to minimize
            clip: (python float) clipping of gradient. If < 0, no clipping
        """
        _lr_m = lr_method.lower() # lower to make sure

        with tf.variable_scope("train_step"):
            if _lr_m == 'adam': # sgd method
                optimizer = tf.train.AdamOptimizer(lr)
            elif _lr_m == 'adagrad':
                optimizer = tf.train.AdagradOptimizer(lr)
            elif _lr_m == 'sgd':
                optimizer = tf.train.GradientDescentOptimizer(lr)
            elif _lr_m == 'rmsprop':
                optimizer = tf.train.RMSPropOptimizer(lr)
            else:
                raise NotImplementedError("Unknown method {}".format(_lr_m))

            if clip > 0: # gradient clipping if clip is positive
                grads, vs     = zip(*optimizer.compute_gradients(loss))
                grads, gnorm  = tf.clip_by_global_norm(grads, clip)
                self.train_op = optimizer.apply_gradients(zip(grads, vs))
            else:
                self.train_op = optimizer.minimize(loss)


    def initialize_session(self):
        """Defines self.sess and initialize the variables"""
        self.logger.info("Initializing tf session")
        self.sess = tf.Session()
        self.sess.run(tf.global_variables_initializer())
        self.saver = tf.train.Saver()


    def restore_session(self, dir_model):
        """Reload weights into session
        Args:
            sess: tf.Session()
            dir_model: dir with weights
        """
        self.logger.info("Reloading the latest trained model...")
        self.saver.restore(self.sess, dir_model)


    def save_session(self):
        """Saves session = weights"""
        if not os.path.exists(self.config.dir_model):
            os.makedirs(self.config.dir_model)
        self.saver.save(self.sess, self.config.dir_model)


    def close_session(self):
        """Closes the session"""
        self.sess.close()


    def add_summary(self):
        """Defines variables for Tensorboard
        Args:
            dir_output: (string) where the results are written
        """
        self.merged = tf.summary.merge_all()
        self.file_writer = tf.summary.FileWriter(self.config.dir_output,
                self.sess.graph)


    def train(self, train, dev):
        """Performs training with early stopping and lr exponential decay
        Args:
            train: dataset that yields tuple of (sentences, tags)
            dev: dataset
        """
        best_score = 0
        nepoch_no_imprv = 0 # for early stopping
        self.add_summary() # tensorboard

        worksheetName = self.config.active_algo + '_' + \
                                 self.config.similarity + '_' + 'split_' + \
                                    str(self.config.split)

        for epoch in range(self.config.nepochs):
            self.logger.info("Epoch {:} out of {:}".format(epoch + 1,
                        self.config.nepochs))

            score = self.run_epoch(train, dev, epoch)
            self.config.lr *= self.config.lr_decay # decay learning rate

            # early stopping and saving best parameters
            if score >= best_score:
                nepoch_no_imprv = 0
                self.save_session()
                best_score = score
                self.logger.info("- new best score!")
            else:
                nepoch_no_imprv += 1
                if nepoch_no_imprv >= self.config.nepoch_no_imprv:
                    self.logger.info("- early stopping {} epochs without "\
                            "improvement".format(nepoch_no_imprv))
                    break

        if self.config.periodic:
            wb = load_workbook(self.config.model + '.xlsx')
            sheet = wb.get_sheet_by_name(worksheetName)
            sheet.cell(row=self.config.excel_id+1, column=1).value = (len(train) + self.config.batch_size - 1) / 149.87
            sheet.cell(row=self.config.excel_id+1, column=2).value = best_score
            wb.save(self.config.model + '.xlsx')
            print('saved excel file')

        elif self.config.mode == 'train':
            wb = xlsxwriter.Workbook(self.config.model + '.xlsx')
            style = wb.add_format({'bold': True, 'font_color': 'red'})
            sheet = wb.get_worksheet_by_name(worksheetName)
            if sheet is None:
                sheet = wb.add_worksheet(worksheetName)
                sheet.write(0, 0, '% samples', style)
                sheet.write(0, 1, 'F1 Score', style)
            sheet.write(self.config.excel_id, 0, (len(train) + self.config.batch_size - 1) / 149.87)
            sheet.write(self.config.excel_id, 1, best_score)
            wb.close()

    def evaluate(self, test, dev, mode):
        """Evaluate model on test set
        Args:
            test: instance of class Dataset
        """
        self.logger.info("Testing model over test set")

        if self.config.mode == 'feedback':
            self.run_evaluate(test, dev, mode)
        else:
            metrics = self.run_evaluate(test, dev, mode)
            msg = " - ".join(["{} {:04.2f}".format(k, v)
                              for k, v in metrics.items()])
            self.logger.info(msg)
            print(msg)
            return metrics


    def get_threshold(self, test, threshold):
        """Evaluate model on test set
        Args:
            test: instance of class Dataset
        """
        self.logger.info("Calculating threshold")
        metrics = self.determine_threshold(test, threshold)
        msg = " - ".join(["{} {:04.2f}".format(k, v)
                for k, v in metrics.items()])
        self.logger.info(msg)
        print(msg)
        return metrics


    def get_encoded(self, sick):
        """
        Encode sentence from SICK data
        Args:
            sick: instance of class SICK dataset
        """
        self.logger.info("\n Encoding sentences of SICK dataset \n")
        self.encode_sents(sick)

